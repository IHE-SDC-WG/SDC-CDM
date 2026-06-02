from __future__ import annotations

import argparse
import json
from collections import Counter
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable

from .json_io import read_json, resolve_repo_path, write_json
from .review_schema import (
    EXPORT_COLUMNS,
    REVIEW_FIELDS,
    REVIEW_STATUSES,
    SOURCE_COLUMNS,
    display_value,
    ensure_review_fields,
    mapping_key,
    normalize_review_field,
    utc_timestamp,
)
from .review_store import DEFAULT_SPEC_PATH, item_mappings, load_spec, mapping_index, save_spec


def _require_openpyxl():
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as error:
        raise RuntimeError(
            "openpyxl is required. Install with: pip install -r phenoml-workflows/requirements.txt"
        ) from error

    return Workbook, load_workbook, Alignment, Font, PatternFill, DataValidation


def _target(mapping: dict[str, Any]) -> str:
    omop_table = str(mapping.get("omop_table") or "").upper()
    mapping_kind = str(mapping.get("mapping_kind") or "").upper()
    if "MEASUREMENT" in omop_table:
        return "MEASUREMENT"
    if "OBSERVATION" in omop_table:
        return "OBSERVATION"
    if mapping_kind == "OMOP_CORE":
        return "OMOP_CORE"
    if mapping.get("proposed_extension_table"):
        return "EXTENSION_TABLE"
    return "UNSPECIFIED"


def _append_table(ws: Any, columns: Iterable[str], rows: Iterable[dict[str, Any]]) -> None:
    ws.append(list(columns))
    for mapping in rows:
        ws.append([display_value(mapping.get(column)) for column in columns])


def _format_table(ws: Any) -> None:
    _, _, Alignment, Font, PatternFill, DataValidation = _require_openpyxl()
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column_cells in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 12), 48)

    if ws.max_row >= 2 and "review_status" in [cell.value for cell in ws[1]]:
        status_col = [cell.value for cell in ws[1]].index("review_status") + 1
        validation = DataValidation(
            type="list",
            formula1=f'"{",".join(REVIEW_STATUSES)}"',
            allow_blank=False,
        )
        ws.add_data_validation(validation)
        validation.add(f"{ws.cell(2, status_col).coordinate}:{ws.cell(ws.max_row, status_col).coordinate}")


def _add_mapping_sheet(wb: Any, title: str, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet(title[:31])
    _append_table(ws, EXPORT_COLUMNS, rows)
    _format_table(ws)


def workbook_bytes(spec: dict[str, Any]) -> bytes:
    Workbook, _, _, Font, _, _ = _require_openpyxl()
    mappings = [ensure_review_fields(dict(mapping)) for mapping in item_mappings(spec)]

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary.append(["Metric", "Value"])
    summary.append(["Total mappings", len(mappings)])
    summary.append(["Generated at", utc_timestamp()])
    summary.append([])
    summary.append(["Review status", "Count"])
    for status, count in Counter(mapping.get("review_status") for mapping in mappings).items():
        summary.append([status, count])
    summary.append([])
    summary.append(["Target", "Count"])
    for target, count in Counter(_target(mapping) for mapping in mappings).items():
        summary.append([target, count])
    for cell in summary[1]:
        cell.font = Font(bold=True)

    _add_mapping_sheet(wb, "All Mappings", mappings)
    _add_mapping_sheet(wb, "MEASUREMENT Candidates", [m for m in mappings if _target(m) == "MEASUREMENT"])
    _add_mapping_sheet(wb, "Extension Table Fields", [m for m in mappings if m.get("proposed_extension_table")])
    _add_mapping_sheet(wb, "OMOP Core Mappings", [m for m in mappings if _target(m) == "OMOP_CORE"])
    _add_mapping_sheet(wb, "Needs Review", [m for m in mappings if m.get("review_status") == "needs_review"])

    person_rows = spec.get("workflow_input", {}).get("naaccr_person_mapping", [])
    ws = wb.create_sheet("NAACCR_PERSON")
    person_columns = list(person_rows[0].keys()) if person_rows else ["field_code", "field_name"]
    _append_table(ws, person_columns, person_rows)
    _format_table(ws)

    fk_rows = spec.get("extension_tables", {}).get("foreign_keys", [])
    ws = wb.create_sheet("FK Relationships")
    fk_columns = list(fk_rows[0].keys()) if fk_rows else ["from_table", "from_column", "to_table", "to_column"]
    _append_table(ws, fk_columns, fk_rows)
    _format_table(ws)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def write_workbook(spec: dict[str, Any], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_bytes(workbook_bytes(spec))


def _sheet_rows(workbook_path: Path) -> list[dict[str, Any]]:
    _, load_workbook, *_ = _require_openpyxl()
    wb = load_workbook(workbook_path, data_only=False)
    if "All Mappings" not in wb.sheetnames:
        raise RuntimeError("Workbook must contain an 'All Mappings' sheet.")
    ws = wb["All Mappings"]
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    rows: list[dict[str, Any]] = []
    for excel_row in ws.iter_rows(min_row=2, values_only=True):
        row = {headers[index]: value for index, value in enumerate(excel_row) if index < len(headers)}
        if not row.get("concept_class_id") and not row.get("concept_code"):
            continue
        row["_excel_row"] = len(rows) + 2
        rows.append(row)
    return rows


def build_import_report(spec: dict[str, Any], workbook_path: Path) -> dict[str, Any]:
    index = mapping_index(spec)
    errors: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []
    changes: list[dict[str, Any]] = []
    ignored_source_edits: list[dict[str, Any]] = []
    seen: set[str] = set()

    for row in _sheet_rows(workbook_path):
        key = f"{row.get('concept_class_id') or ''}::{row.get('concept_code') or ''}"
        excel_row = row["_excel_row"]
        mapping = index.get(key)
        if mapping is None:
            errors.append({"row": excel_row, "key": key, "message": "Unknown mapping key."})
            continue
        if key in seen:
            errors.append({"row": excel_row, "key": key, "message": "Duplicate mapping key."})
            continue
        seen.add(key)

        for column in SOURCE_COLUMNS:
            workbook_value = display_value(row.get(column))
            canonical_value = display_value(mapping.get(column))
            if workbook_value != canonical_value:
                ignored_source_edits.append(
                    {
                        "row": excel_row,
                        "key": key,
                        "field": column,
                        "workbook_value": workbook_value,
                        "canonical_value": canonical_value,
                    }
                )

        for field in REVIEW_FIELDS:
            if field not in row:
                warnings.append({"row": excel_row, "key": key, "message": f"Missing review field: {field}"})
                continue
            new_value = normalize_review_field(field, row.get(field))
            if field == "review_status" and new_value not in REVIEW_STATUSES:
                errors.append(
                    {
                        "row": excel_row,
                        "key": key,
                        "field": field,
                        "message": f"Invalid review_status: {new_value}",
                    }
                )
                continue
            old_value = normalize_review_field(field, mapping.get(field))
            if new_value != old_value:
                changes.append(
                    {
                        "row": excel_row,
                        "key": key,
                        "concept_class_id": mapping.get("concept_class_id"),
                        "concept_code": mapping.get("concept_code"),
                        "concept_name": mapping.get("concept_name"),
                        "field": field,
                        "old_value": old_value,
                        "new_value": new_value,
                    }
                )

    return {
        "valid": not errors,
        "generated_at": utc_timestamp(),
        "changes": changes,
        "ignored_source_edits": ignored_source_edits,
        "errors": errors,
        "warnings": warnings,
        "summary": {
            "changes": len(changes),
            "ignored_source_edits": len(ignored_source_edits),
            "errors": len(errors),
            "warnings": len(warnings),
        },
    }


def apply_import_report(spec: dict[str, Any], report: dict[str, Any]) -> dict[str, Any]:
    if not report.get("valid"):
        raise RuntimeError("Cannot apply invalid import report.")

    index = mapping_index(spec)
    applied = 0
    for change in report.get("changes", []):
        mapping = index.get(change["key"])
        if mapping is None:
            raise RuntimeError(f"Unknown mapping key during apply: {change['key']}")
        mapping[change["field"]] = change["new_value"]
        applied += 1

    report = dict(report)
    report["applied"] = True
    report["applied_at"] = utc_timestamp()
    report["summary"] = dict(report.get("summary", {}))
    report["summary"]["applied_changes"] = applied
    return report


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export/import NAACCR OMOP review Excel.")
    subparsers = parser.add_subparsers(dest="command", required=True)

    export = subparsers.add_parser("export")
    export.add_argument("--spec", default=DEFAULT_SPEC_PATH)
    export.add_argument("--output", required=True, type=Path)

    import_cmd = subparsers.add_parser("import")
    import_cmd.add_argument("--spec", default=DEFAULT_SPEC_PATH)
    import_cmd.add_argument("--input", required=True, type=Path)
    import_cmd.add_argument("--patch-report", required=True, type=Path)
    import_cmd.add_argument("--apply", action="store_true")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    spec = load_spec(args.spec)

    if args.command == "export":
        write_workbook(spec, args.output)
        print(f"Wrote {args.output}")
        return 0

    report = build_import_report(spec, args.input)
    if args.apply and report["valid"]:
        report = apply_import_report(spec, report)
        save_spec(spec, args.spec)
    write_json(args.patch_report, report)
    print(json.dumps(report["summary"], indent=2))
    return 0 if report["valid"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
